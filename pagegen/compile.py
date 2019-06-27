#!/usr/bin/env bash

'''
------
Auto-compile Webpages for ICAPS 2019
------
'''

'''
packages
'''

import openpyxl as xl
import argparse, sys, random, copy

'''
global variables 
'''

data = {} 
pc_list = {}
paper_list = {}
demos_list = {}
jorunal_list = {}
program = {}

'''
html templates
'''

with open('templates/index-template.html', 'r') as temp:
    index_template = temp.read()

with open('templates/cfp-template.html', 'r') as temp:
    cfp_template = temp.read()

with open('templates/workshop-template.html', 'r') as temp:
    workshop_template = temp.read()

with open('templates/tutorial-template.html', 'r') as temp:
    tutorial_template = temp.read()

with open('templates/info-template.html', 'r') as temp:
    info_template = temp.read()

with open('templates/awards-template.html', 'r') as temp:
    awards_template = temp.read()

with open('templates/header-template.html', 'r') as temp:
    header_template = temp.read()

with open('templates/navbar-template.html', 'r') as temp:
    navbar_template = temp.read()

with open('templates/quicklinks-template.html', 'r') as temp:
    quicklinks_template = temp.read()

with open('templates/organizing-team-template.html', 'r') as temp:
    organizing_team_template = temp.read()

with open('templates/banner-template.html', 'r') as temp:
    banner_template = temp.read()

with open('templates/pc-info-template.html', 'r') as temp:
    track_table_td = temp.read()

with open('templates/privacy-policy-template.html', 'r') as temp:
    privacy_policy_template = temp.read()

with open('templates/terms-of-use-template.html', 'r') as temp:
    terms_of_use_template = temp.read()

'''
program templates
'''

with open('templates/program-template.html', 'r') as temp:
    program_template = temp.read()

with open('templates/accepted-papers-template.html', 'r') as temp:
    accepted_papers_template = temp.read()

with open('templates/demo-template.html', 'r') as temp:
    demo_template = temp.read()

with open('templates/demo-program-entry.html', 'r') as temp:
    demo_program_entry = temp.read()

with open('templates/journal-track-template.html', 'r') as temp:
    journal_track_template = temp.read()

with open('templates/invited-talks-template.html', 'r') as temp:
    invited_talks_template = temp.read()

with open('templates/track-paper-single.html', 'r') as temp:
    track_paper_single_template = temp.read()

with open('templates/track-table.html', 'r') as temp:
    track_table_template = temp.read()

with open('templates/track-table-single.html', 'r') as temp:
    track_table_single_template = temp.read()

with open('templates/paper-info-template.html', 'r') as temp:
    paper_info_stub = temp.read()

with open('templates/demo-journal-info-template.html', 'r') as temp:
    demo_journal_info_stub = temp.read()

'''
program details templates
'''

with open('templates/program-details-template.html', 'r') as temp:
    program_details_template = temp.read()

with open('templates/program/program-td-event.html', 'r') as temp:
    program_details_event_template = temp.read()

with open('templates/program/program-td-invited-talk.html', 'r') as temp:
    program_details_invited_talk_template = temp.read()

with open('templates/program/program-td-session.html', 'r') as temp:
    program_details_session_template = temp.read()

with open('templates/program/program-td-paper.html', 'r') as temp:
    program_details_paper_template = temp.read()

'''
carousel templates
'''

with open('templates/carousel-elements/carousel-template.html', 'r') as temp:
    carousel_template = temp.read()

with open('templates/carousel-elements/carousel-entry-template.html', 'r') as temp:
    carousel_entry_template = temp.read()

with open('templates/carousel-elements/carousel-entry-link-template.html', 'r') as temp:
    carousel_entry_link_template = temp.read()

with open('templates/carousel-elements/carousel-inner-template.html', 'r') as temp:
    carousel_inner_template = temp.read()

with open('templates/carousel-elements/carousel-indicators-template.html', 'r') as temp:
    carousel_indicators_template = temp.read()

'''
method :: cache data from xlsx file
'''
def cache(filename = 'data.xlsx', 
    pc_filename = 'icaps19_info/ICAPS-2019_PC.xlsx', 
    papers_filename = 'icaps19_info/ICAPS19 Metadata.xlsx', 
    demos_filename='icaps19_info/demo.xlsx',
    journals_filename='icaps19_info/journal_track.xlsx',
    program_filename='icaps19_info/ICAPS19 Schedule.xlsx'):

    global data, pc_list, paper_list, demos_list, jorunal_list

    print( 'Reading highlights...' )

    wb = xl.load_workbook(filename)

    count = 0

    for item in wb['items']:

        new_entry = [str(elem.value) for elem in item]

        if not count: keys = new_entry
        else:

            data[count] = {}
            inner_count  = 0

            for key in keys:

                data[count][key] = new_entry[inner_count]
                inner_count += 1

        count += 1

    print( 'Reading PC list...' )

    wb = xl.load_workbook(pc_filename)

    for sheet_name in wb.sheetnames:

        dict_entry = {}

        for row in wb[sheet_name]:

            row_values = [str(item.value).strip() for item in row][:4]

            key   = row_values[3]
            entry = ["{} {}".format(row_values[0], row_values[1]), row_values[2]]

            if key in dict_entry:
                dict_entry[key].append(entry)
            else:
                dict_entry[key] = [entry]

        pc_list[str(sheet_name)] = dict_entry


    print( 'Reading Paper list...' )

    wb = xl.load_workbook(papers_filename)

    title_flag = False
    for row in wb["AAAIPressList"]:

        if not title_flag:
            title_flag = True
        else:

            row_values = [str(item.value).strip() for item in row]

            key   = row_values[1]
            entry = row_values

            if key in paper_list:
                paper_list[key].append(entry)
            else:
                paper_list[key] = [entry]

    print( 'Reading Demos list...' )

    wb = xl.load_workbook(demos_filename)

    for row in wb["DEMOS"]:

        row_values = [str(item.value).strip() for item in row]

        key   = row_values[1]
        entry = row_values

        demos_list[key] = entry

    print( 'Journal Track list...' )

    wb = xl.load_workbook(journals_filename)

    for row in wb["JOURNALS"]:

        row_values = [str(item.value).strip() for item in row]

        key   = row_values[1]
        entry = row_values

        jorunal_list[key] = entry

    print( 'Program list...' )

    wb = xl.load_workbook(program_filename)

    current_date_key = ""
    current_time_key = ""

    for row in wb["Details"]:

        row_values = [str(item.value).strip() for item in row]

        if not all(item == 'None' for item in row_values):

            if 'th' in row_values[0]:

                current_date_key = row_values[0]
                program[current_date_key] = {}
                continue

            else:

                if row_values[0] != 'None':
                    current_time_key = row_values[0]
                    program[current_date_key][current_time_key] = []

                program[current_date_key][current_time_key].append(row_values[1:])


'''
method :: write index.html
'''
def write_file(args):

    global index_template, cfp_template, workshop_template, organizing_team_template, tutorial_template, info_template, privacy_policy_template, terms_of_use_template, awards_template
    global program_template, program_details_template, invited_talks_template, accepted_papers_template, journal_track_template
    global program_details_template, program_details_event_template, program_details_invited_talk_template, program_details_session_template, program_details_paper_template
    global demo_program_entry, demo_journal_info_stub, demo_template

    # cache data
    print( 'Reading data...' )
    cache()

    # write problem file
    print( 'Compiling index.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    index_template = index_template.replace('[HEADER]', header_template)    
    index_template = index_template.replace('[NAVBAR]', navbar_template)    
    index_template = index_template.replace('[BANNER]', banner_template)    
    index_template = index_template.replace('[QUICKLINKS]', quicklinks_template)    

    # write primary carousel section
    print( 'Writing carousel ...' )

    carousel_indicators = ''
    carousel_inner = ''
    carousel_entries = ''

    for key in data:

        carousel_indicator = copy.deepcopy(carousel_indicators_template) 
        carousel_indicator = carousel_indicator.replace('[NUMBER]', str(key))

        if key == 1:
            carousel_indicator = carousel_indicator.replace('class', 'class="active"')

        carousel_indicators += carousel_indicator

        local_carousel_inner = copy.deepcopy(carousel_inner_template)

        local_carousel_inner = local_carousel_inner.replace('[NUMBER]', str(key)) 
        local_carousel_inner = local_carousel_inner.replace('[Name]', data[key]['Name']) 
        local_carousel_inner = local_carousel_inner.replace('[Image]', data[key]['Image']) 

        carousel_inner += '\n\n' + local_carousel_inner

        carousel_entry = copy.deepcopy(carousel_entry_template) 
        carousel_entry = carousel_entry.replace('[NUMBER]', str(key))

        for inner_key in data[key]:

            new_entry = ""

            if inner_key == 'Links':

                link_list = data[key][inner_key].split(',')

                for link in link_list:

                    link_entry = copy.deepcopy(carousel_entry_link_template)
                    link_entry = link_entry.replace('[Link]', link)

                    new_entry += '\n' + link_entry

            else: new_entry = data[key][inner_key]

            carousel_entry = carousel_entry.replace('[{}]'.format(inner_key), new_entry)                

        carousel_entries += '\n\n' + carousel_entry

    local_carousel_template = copy.deepcopy(carousel_template)
    local_carousel_template = local_carousel_template.replace('[CAROUSEL-INDICATORS]', carousel_indicators)
    local_carousel_template = local_carousel_template.replace('[CAROUSEL-INNER]', carousel_inner)
    local_carousel_template = local_carousel_template.replace('[CAROUSEL-ELEMENTS]', carousel_entries)

    index_template = index_template.replace('[CAROUSEL]', local_carousel_template)

    # write to output
    print( 'Writing to file (index.html) ...' )

    with open('../index.html', 'w') as output_file:
        output_file.write(index_template)

    # write cfp file
    print( 'Compiling calls.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    cfp_template = cfp_template.replace('[HEADER]', header_template)    
    cfp_template = cfp_template.replace('[NAVBAR]', navbar_template)    
    cfp_template = cfp_template.replace('[BANNER]', banner_template)    
    cfp_template = cfp_template.replace('[QUICKLINKS]', quicklinks_template)    

    # write to output
    print( 'Writing to file (calls.html) ...' )

    with open('../calls.html', 'w') as output_file:
        output_file.write(cfp_template)

    # write tutorials file
    print( 'Compiling tutorials.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    tutorial_template = tutorial_template.replace('[HEADER]', header_template)    
    tutorial_template = tutorial_template.replace('[NAVBAR]', navbar_template)    
    tutorial_template = tutorial_template.replace('[BANNER]', banner_template)    
    tutorial_template = tutorial_template.replace('[QUICKLINKS]', quicklinks_template)    

    # write to output
    print( 'Writing to file (tutorials.html) ...' )

    with open('../tutorials.html', 'w') as output_file:
        output_file.write(tutorial_template)

    # write info file
    print( 'Compiling info.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    info_template = info_template.replace('[HEADER]', header_template)    
    info_template = info_template.replace('[NAVBAR]', navbar_template)    
    info_template = info_template.replace('[BANNER]', banner_template)    
    info_template = info_template.replace('[QUICKLINKS]', quicklinks_template)    

    # write to output
    print( 'Writing to file (info.html) ...' )

    with open('../info.html', 'w') as output_file:
        output_file.write(info_template)

    # write info file
    print( 'Compiling awards.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    awards_template = awards_template.replace('[HEADER]', header_template)    
    awards_template = awards_template.replace('[NAVBAR]', navbar_template)    
    awards_template = awards_template.replace('[BANNER]', banner_template)    
    awards_template = awards_template.replace('[QUICKLINKS]', quicklinks_template)    

    # write to output
    print( 'Writing to file (awards.html) ...' )

    with open('../awards.html', 'w') as output_file:
        output_file.write(awards_template)

    # write workshops file
    print( 'Compiling workshops.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    workshop_template = workshop_template.replace('[HEADER]', header_template)    
    workshop_template = workshop_template.replace('[NAVBAR]', navbar_template)    
    workshop_template = workshop_template.replace('[BANNER]', banner_template)    
    workshop_template = workshop_template.replace('[QUICKLINKS]', quicklinks_template)    

    # write to output
    print( 'Writing to file (workshops.html) ...' )

    with open('../workshops.html', 'w') as output_file:
        output_file.write(workshop_template)

    # write committees file
    print( 'Compiling organizing_team.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    track_tables_stub = ""

    for track in pc_list:

        temp_track_table_template = copy.deepcopy(track_table_template).replace('[TRACK-NAME]', "{} Track".format(track))
        temp_table_stub = ""

        for role in pc_list[track]:

            temp_table_stub_single = ""

            for person in pc_list[track][role]:

                if person[1] != "None":
                    temp_table_stub_single += track_table_td.replace('[NAME]', person[0]).replace('[AFFILIATION]', person[1]).replace("d-none", "")
                else:
                    temp_table_stub_single += track_table_td.replace('[NAME]', person[0]).replace('[AFFILIATION]', "")

            temp_track_table_single_template = copy.deepcopy(track_table_single_template).replace('[ROLE]', role)
            temp_track_table_single_template = temp_track_table_single_template.replace('[ENTRIES]', temp_table_stub_single)

            temp_table_stub += "\n\n" + temp_track_table_single_template

        temp_track_table_template = temp_track_table_template.replace('[TRACK-TABLE]', temp_table_stub)
        track_tables_stub += "\n\n\n" + temp_track_table_template

    organizing_team_template = organizing_team_template.replace('[HEADER]', header_template)    
    organizing_team_template = organizing_team_template.replace('[NAVBAR]', navbar_template)    
    organizing_team_template = organizing_team_template.replace('[BANNER]', banner_template)    
    organizing_team_template = organizing_team_template.replace('[QUICKLINKS]', quicklinks_template)    

    organizing_team_template = organizing_team_template.replace('[TRACK-TABLES]', track_tables_stub)

    # write to output
    print( 'Writing to file (organizing-team.html) ...' )

    with open('../organizing-team.html', 'wb') as output_file:
        output_file.write(organizing_team_template.encode("utf-8"))

    # write program file
    print( 'Compiling program.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    program_template = program_template.replace('[HEADER]', header_template)    
    program_template = program_template.replace('[NAVBAR]', navbar_template)    
    program_template = program_template.replace('[BANNER]', banner_template)    
    program_template = program_template.replace('[QUICKLINKS]', quicklinks_template)    

    # write to output
    print( 'Writing to file (program.html) ...' )

    with open('../program.html', 'wb') as output_file:
        output_file.write(program_template.encode("utf-8"))

    # write program file
    print( 'Compiling program-details.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    program_details_template = program_details_template.replace('[HEADER]', header_template)    
    program_details_template = program_details_template.replace('[NAVBAR]', navbar_template)    
    program_details_template = program_details_template.replace('[BANNER]', banner_template)    
    program_details_template = program_details_template.replace('[QUICKLINKS]', quicklinks_template)    

    program_stub = ""
    for date in program:
        row_num = 0
        for time in program[date]:
            row_num += 1
            row_num_2 = 1 

            if 'invited talk' in program[date][time][0][0].lower():
                program_stub += program_details_invited_talk_template.replace('[DATE]', date).replace('[TIME]', time).replace('[TALK]', program[date][time][0][0])
            
            elif program[date][time][0][1] != 'None':
                program_stub += program_details_session_template.replace('[TIME]', time).replace('[SESSION-1]', program[date][time][0][0]).replace('[SESSION-2]', program[date][time][0][3])

                for session in program[date][time][1:]:
                    row_num += 1
                    row_num_2 += 1

                    paper1_details = '{} &middot; <span class="text-muted">{}</span>'.format(session[0], session[1])

                    # HACK ¯\_(ツ)_/¯
                    if "a:" in session[0] or "b:" in session[0] or "a:" in session[3] or "b:" in session[3]:
                        paper2_details = session[3]
                    else:
                        paper2_details = '{} &middot; <span class="text-muted">{}</span>'.format(session[3], session[4])

                    temp = program_details_paper_template.replace('[SESSION-1]', paper1_details).replace('[SESSION-2]', paper2_details)

                    if session[2] != 'None':
                        temp = temp.replace('d-none-1 d-none', '').replace('[EXTRA-1]', session[2])

                    if session[5] != 'None':
                        temp = temp.replace('d-none-2 d-none', '').replace('[EXTRA-2]', session[5])

                    if session[2] == 'Short Paper' or session[2] == 'Journal Paper':
                        temp = temp.replace('text-orange-1', '')
                    else:
                        temp = temp.replace('text-orange-1', 'text-orange')

                    if session[5] == 'Short Paper' or session[5] == 'Journal Paper':
                        temp = temp.replace('text-orange-2', '')
                    else:
                        temp = temp.replace('text-orange-2', 'text-orange')

                    # HACK ¯\_(ツ)_/¯
                    if "a:" in session[0] or "b:" in session[0] or "a:" in session[3] or "b:" in session[3]:
                        pass
                    else:
                        temp = temp.replace('<b>', '').replace('</b>', '')

                    program_stub += temp

            else:
                program_stub += program_details_event_template.replace('[TIME]', time).replace('[EVENT]', program[date][time][0][0])

            program_stub = program_stub.replace('[ROWS-2]', str(row_num_2))

        program_stub = program_stub.replace('[ROWS]', str(row_num))

    program_details_template = program_details_template.replace('[PROGRAM]', program_stub).replace('None', '')

    # write to output
    print( 'Writing to file (program-details.html) ...' )

    with open('../program-details.html', 'wb') as output_file:
        output_file.write(program_details_template.encode("utf-8"))

    # write program file
    print( 'Compiling invited-talks.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    invited_talks_template = invited_talks_template.replace('[HEADER]', header_template)    
    invited_talks_template = invited_talks_template.replace('[NAVBAR]', navbar_template)    
    invited_talks_template = invited_talks_template.replace('[BANNER]', banner_template)    
    invited_talks_template = invited_talks_template.replace('[QUICKLINKS]', quicklinks_template)    

    # write to output
    print( 'Writing to file (invited-talks.html) ...' )

    with open('../invited-talks.html', 'wb') as output_file:
        output_file.write(invited_talks_template.encode("utf-8"))

    # write program file
    print( 'Compiling accepted-papers.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    paper_list_stub = ""

    track_order = ["Main", "Applications", "Planning & Learning", "Robotics"]

    for track in track_order:

        temp_track_paper_single_template = copy.deepcopy(track_paper_single_template).replace('[TRACK]', track + ' Track')

        temp_papers_list_stub = ""

        for paper in paper_list[track]:

            authors = []
            for i in range(5, len(paper)-1, 3):
                if paper[i].strip() != "" and paper[i].strip() != "None":
                    authors.append(paper[i])

            temp = paper_info_stub.replace('[TITLE]', paper[2]).replace('[AUTHORS]', ', '.join(authors))

            if paper[-1] != "None":
                temp = temp.replace('[AWARD]', paper[-1])
            else:
                temp = temp.replace('toggle-', '')

            # save paper info

            authors = []
            for i in range(5, len(paper)-1, 3):
                if paper[i].strip() != "" and paper[i].strip() != "None":

                    if paper[i+1] == "None":
                        authors.append('<span class="profile">{}</span>'.format(paper[i]))
                    else:
                        authors.append('<span class="profile">{}</span> ({})'.format(paper[i], paper[i+1]))

            temp = temp.replace('[save-paper-title]', '<strong>{}</strong>'.format(paper[2]))
            temp = temp.replace('[save-paper-authors]', ', '.join(authors))
            temp = temp.replace('[save-paper-abstract]', paper[3])
            temp = temp.replace('[save-paper-contact]', paper[4][::-1])

            temp_papers_list_stub += temp

        paper_list_stub += "\n\n" + temp_track_paper_single_template.replace('[ENTRIES]', temp_papers_list_stub)

    accepted_papers_template = accepted_papers_template.replace('[HEADER]', header_template)    
    accepted_papers_template = accepted_papers_template.replace('[NAVBAR]', navbar_template)    
    accepted_papers_template = accepted_papers_template.replace('[BANNER]', banner_template)    
    accepted_papers_template = accepted_papers_template.replace('[QUICKLINKS]', quicklinks_template)    

    accepted_papers_template = accepted_papers_template.replace('[PAPERS]', paper_list_stub)

    # write to output
    print( 'Writing to file (accepted-papers.html) ...' )

    with open('../accepted-papers.html', 'wb') as output_file:
        output_file.write(accepted_papers_template.encode("utf-8"))

    # write program file
    print( 'Compiling demos.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    demos_list_stub = ""
    for demo in demos_list:

        if demos_list[demo][2].strip() == 'None':
            demo_link =  ""
        else:
            demo_link = 'href="{}" target="_blank"'.format(demos_list[demo][2].strip())

        demos_list_stub += demo_journal_info_stub.replace('[TITLE]', demo).replace('[AUTHORS]', demos_list[demo][0]).replace('[LINK]', demo_link)

    demos_program_stub = ""
    num_desks = max([int(demos_list[demo][4]) for demo in demos_list])
    num_session = set([demos_list[demo][3] for demo in demos_list])

    for desk in range(num_desks):
        demo_program_entry_stub = copy.deepcopy(demo_program_entry)

        for session in num_session:

            for demo in demos_list:

                if demos_list[demo][3] == session and demos_list[demo][4] == str(desk+1):

                    if demos_list[demo][2].strip() == 'None':
                        demo_link =  ""
                    else:
                        demo_link = 'href="{}" target="_blank"'.format(demos_list[demo][2].strip())

                    temp = '<a class="text-dark" [LINK]>[TITLE]<span class="text-muted"> &bull; [AUTHORS]</span></a>'
                    temp = temp.replace('[TITLE]', demo).replace('[AUTHORS]', demos_list[demo][0]).replace('[LINK]', demo_link)

                    demo_program_entry_stub = demo_program_entry_stub.replace('[NUM]', str(desk+1)).replace('[SESSION-{}]'.format(session), temp).replace('[LINK-{}]'.format(session), demo_link)

                    if demo_link:
                        demo_program_entry_stub = demo_program_entry_stub.replace('demo-paper-link-{}'.format(session), 'demo-paper-link click-to-go')

        demo_program_entry_stub = demo_program_entry_stub.replace('[SESSION-A]', '').replace('[SESSION-B]', '')
        demos_program_stub += demo_program_entry_stub

    demo_template = demo_template.replace('[HEADER]', header_template)    
    demo_template = demo_template.replace('[NAVBAR]', navbar_template)    
    demo_template = demo_template.replace('[BANNER]', banner_template)    
    demo_template = demo_template.replace('[QUICKLINKS]', quicklinks_template)    

    demo_template = demo_template.replace('[DEMO-PROGRAM]', demos_program_stub)
    demo_template = demo_template.replace('[DEMOS]', demos_list_stub)

    # write to output
    print( 'Writing to file (demos.html) ...' )

    with open('../demos.html', 'wb') as output_file:
        output_file.write(demo_template.encode("utf-8"))

    # write program file
    print( 'Compiling journal-track.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    journal_list_stub = ""
    for journal in jorunal_list:

        journal_link =  ""
        if len(jorunal_list[journal]) == 3:
            if jorunal_list[journal][2].strip() != 'None':
                journal_link = 'href="{}" target="_blank"'.format(jorunal_list[journal][2].strip())

        journal_list_stub += demo_journal_info_stub.replace('[TITLE]', journal).replace('[AUTHORS]', jorunal_list[journal][0]).replace('[LINK]', journal_link)

    journal_track_template = journal_track_template.replace('[HEADER]', header_template)    
    journal_track_template = journal_track_template.replace('[NAVBAR]', navbar_template)    
    journal_track_template = journal_track_template.replace('[BANNER]', banner_template)    
    journal_track_template = journal_track_template.replace('[QUICKLINKS]', quicklinks_template)    

    journal_track_template = journal_track_template.replace('[DEMOS]', journal_list_stub)

    # write to output
    print( 'Writing to file (journal-track.html) ...' )

    with open('../journal-track.html', 'wb') as output_file:
        output_file.write(journal_track_template.encode("utf-8"))

    # write privacy policy file
    print( 'Compiling privacy-policy.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    privacy_policy_template = privacy_policy_template.replace('[HEADER]', header_template)    
    privacy_policy_template = privacy_policy_template.replace('[NAVBAR]', navbar_template)    
    privacy_policy_template = privacy_policy_template.replace('[BANNER]', banner_template)    

    # write to output
    print( 'Writing to file (privacy-policy.html) ...' )

    with open('../privacy-policy.html', 'w') as output_file:
        output_file.write(privacy_policy_template)

    # write terms of use file
    print( 'Compiling terms-of-use.html ...' )

    # writing templates
    print( 'Writing templates ...' )

    terms_of_use_template = terms_of_use_template.replace('[HEADER]', header_template)    
    terms_of_use_template = terms_of_use_template.replace('[NAVBAR]', navbar_template)    
    terms_of_use_template = terms_of_use_template.replace('[BANNER]', banner_template)    

    # write to output
    print( 'Writing to file (terms-of-use.html) ...' )

    with open('../terms-of-use.html', 'w') as output_file:
        output_file.write(terms_of_use_template)

    print( 'Done.' )


'''
method :: main
'''
def main():

    parser = argparse.ArgumentParser(description='''Auto-compile Webpages of ICAPS 2019 Homepage.''', epilog='''Usage >> python compile_carousel.py''')
    args = parser.parse_args()

    if '-h' in sys.argv[1:]:
        print( parser.print_help() )
        sys.exit(1)
    else:
        write_file(args)


if __name__ == "__main__":
    main()
